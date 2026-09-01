import io
import re
import unicodedata
from datetime import datetime, time, timedelta

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


st.set_page_config(
    page_title="Análise de Setups de Matriz",
    page_icon="⚙️",
    layout="wide",
    initial_sidebar_state="expanded",
)


CORES = {
    "azul": "#2563EB",
    "verde": "#10B981",
    "laranja": "#F59E0B",
    "vermelho": "#EF4444",
    "roxo_claro": "#A78BFA",
    "roxo_escuro": "#6D28D9",
}


CLASSIFICACOES_PRINCIPAL = [
    "Sem simultaneidade",
    "Simultâneo dentro da capacidade",
    "Simultâneo fora do intervalo de excesso",
    "Setup crítico por excesso de capacidade",
    "Dado não analisável",
]


COLUNAS_BASE = [
    "id_analise",
    "setup_id",
    "extra_id",
    "tipo_setup",
    "equipamento",
    "setor",
    "turno",
    "linha_inicial",
    "linha_final",
    "inicio_setup",
    "fim_setup",
    "tempo_total_segundos",
    "tempo_total_formatado",
    "origem_tempo_total",
    "tempo_total_relatorio_segundos",
    "soma_tempo_linhas_segundos",
    "tempo_total_datas_segundos",
    "inicio_aguardando",
    "fim_aguardando",
    "inicio_troca_molde",
    "fim_troca_molde",
    "aguardando_segundos",
    "aguardando_formatado",
    "troca_molde_segundos",
    "troca_molde_formatada",
    "inicio_janela",
    "fim_janela",
    "tempo_janela_segundos",
    "janela_simultaneidade_formatada",
    "descricoes_bloco",
    "descricao",
    "codigo",
    "programa",
    "documento",
    "produto",
    "ferramental",
    "qualidade",
    "participa_analise",
    "setup_extra",
    "setup_simultaneo",
    "bloco_id",
    "bloco_critico",
    "setup_critico",
    "pico_maquinas_paradas",
    "excedente_maximo",
    "tempo_excesso_segundos",
    "tempo_excesso_formatado",
    "tempo_critico_individual_segundos",
    "tempo_critico_individual_formatado",
    "espera_coincidente_excesso_segundos",
    "espera_coincidente_excesso_formatada",
    "quantidade_criticos_bloco",
    "equipamentos_criticos_bloco",
    "classificacao",
    "justificativa",
    "motivo_extra",
    "fonte_setup",
]


def normalizar_texto(valor):
    if valor is None:
        return ""

    try:
        if pd.isna(valor):
            return ""
    except (TypeError, ValueError):
        pass

    texto = unicodedata.normalize(
        "NFKD",
        str(valor),
    )

    texto = texto.encode(
        "ascii",
        "ignore",
    ).decode(
        "ascii",
    )

    texto = texto.strip().lower()
    texto = re.sub(
        r"\s+",
        " ",
        texto,
    )

    return texto


def normalizar_coluna(valor):
    texto = normalizar_texto(
        valor
    )

    texto = texto.replace(
        " ",
        "_",
    )

    texto = re.sub(
        r"[^a-z0-9_]",
        "",
        texto,
    )

    texto = re.sub(
        r"_+",
        "_",
        texto,
    )

    return texto.strip("_")


def normalizar_equipamento(valor):
    if valor is None:
        return ""

    texto = str(valor)

    texto = re.sub(
        r"\s+",
        "",
        texto,
    )

    return texto.upper()


def localizar_coluna(
    dataframe,
    alternativas,
    obrigatoria=False,
):
    mapa_colunas = {
        normalizar_coluna(coluna): coluna
        for coluna in dataframe.columns
    }

    for alternativa in alternativas:
        chave = normalizar_coluna(
            alternativa
        )

        if chave in mapa_colunas:
            return mapa_colunas[chave]

    if obrigatoria:
        raise ValueError(
            "Coluna obrigatória não encontrada: "
            + ", ".join(alternativas)
        )

    return None


def preparar_dataframe(dataframe):
    saida = dataframe.copy()

    saida.columns = [
        normalizar_coluna(coluna)
        for coluna in saida.columns
    ]

    saida = saida.dropna(
        how="all",
    )

    saida = saida.reset_index(
        drop=True,
    )

    return saida


def duracao_segundos(valor):
    if valor is None:
        return np.nan

    try:
        if pd.isna(valor):
            return np.nan
    except (TypeError, ValueError):
        pass

    if isinstance(
        valor,
        (
            pd.Timedelta,
            timedelta,
        ),
    ):
        return max(
            0.0,
            float(
                valor.total_seconds()
            ),
        )

    if isinstance(
        valor,
        time,
    ):
        segundos = (
            valor.hour * 3600
            + valor.minute * 60
            + valor.second
            + valor.microsecond / 1_000_000
        )

        return float(segundos)

    if isinstance(
        valor,
        (
            int,
            float,
            np.integer,
            np.floating,
        ),
    ):
        numero = float(valor)

        if numero < 0:
            return np.nan

        if 0 < numero < 1:
            return numero * 86400.0

        return numero

    texto = str(valor).strip()
    texto = texto.replace(
        ",",
        ".",
    )

    if not texto:
        return np.nan

    if texto.lower() in {
        "nan",
        "nat",
        "none",
        "null",
    }:
        return np.nan

    if re.fullmatch(
        r"\d+(\.\d+)?",
        texto,
    ):
        numero = float(texto)

        if 0 < numero < 1:
            return numero * 86400.0

        return numero

    partes = texto.split(":")

    try:
        if len(partes) == 3:
            horas = float(
                partes[0]
            )

            minutos = float(
                partes[1]
            )

            segundos = float(
                partes[2]
            )

            return max(
                0.0,
                horas * 3600
                + minutos * 60
                + segundos,
            )

        if len(partes) == 2:
            minutos = float(
                partes[0]
            )

            segundos = float(
                partes[1]
            )

            return max(
                0.0,
                minutos * 60
                + segundos,
            )

    except ValueError:
        return np.nan

    try:
        timedelta_convertido = (
            pd.to_timedelta(texto)
        )

        return max(
            0.0,
            float(
                timedelta_convertido.total_seconds()
            ),
        )

    except Exception:
        return np.nan


def formatar_duracao(segundos):
    try:
        if (
            segundos is None
            or pd.isna(segundos)
        ):
            return "Não disponível"
    except (TypeError, ValueError):
        return "Não disponível"

    total = max(
        0,
        int(
            round(
                float(segundos)
            )
        ),
    )

    horas, restante = divmod(
        total,
        3600,
    )

    minutos, segundos_finais = divmod(
        restante,
        60,
    )

    return (
        f"{horas:02d}:"
        f"{minutos:02d}:"
        f"{segundos_finais:02d}"
    )


def soma_duracoes(serie):
    if (
        serie is None
        or len(serie) == 0
    ):
        return np.nan

    valores = serie.apply(
        duracao_segundos
    )

    valores = valores.dropna()

    if len(valores) == 0:
        return np.nan

    return float(
        valores.sum()
    )


def converter_datas(serie):
    convertido = pd.to_datetime(
        serie,
        errors="coerce",
        dayfirst=False,
    )

    if (
        len(convertido) > 0
        and convertido.notna().mean() < 0.5
    ):
        convertido = pd.to_datetime(
            serie,
            errors="coerce",
            dayfirst=True,
        )

    return convertido


def ler_arquivo(arquivo):
    nome_arquivo = str(
        getattr(
            arquivo,
            "name",
            "",
        )
    ).lower()

    arquivo.seek(0)
    conteudo = arquivo.read()
    arquivo.seek(0)

    if not conteudo:
        raise ValueError(
            "O arquivo enviado está vazio."
        )

    if nome_arquivo.endswith(
        (
            ".csv",
            ".txt",
        )
    ):
        ultimo_erro = None

        for codificacao in [
            "utf-8-sig",
            "utf-8",
            "cp1252",
            "latin-1",
        ]:
            try:
                return pd.read_csv(
                    io.BytesIO(
                        conteudo
                    ),
                    sep=";",
                    encoding=codificacao,
                    dtype=object,
                    low_memory=False,
                )

            except Exception as erro:
                ultimo_erro = erro

        raise ValueError(
            "Não foi possível ler o CSV: "
            f"{ultimo_erro}"
        )

    assinatura_xlsx = (
        conteudo[:4]
        == b"PK\x03\x04"
    )

    assinatura_xls = (
        conteudo[:8]
        == b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"
    )

    if assinatura_xlsx:
        return pd.read_excel(
            io.BytesIO(
                conteudo
            ),
            engine="openpyxl",
            dtype=object,
        )

    if assinatura_xls:
        return pd.read_excel(
            io.BytesIO(
                conteudo
            ),
            engine="xlrd",
            dtype=object,
        )

    if nome_arquivo.endswith(
        (
            ".xlsx",
            ".xlsm",
        )
    ):
        return pd.read_excel(
            io.BytesIO(
                conteudo
            ),
            engine="openpyxl",
            dtype=object,
        )

    if nome_arquivo.endswith(
        ".xls"
    ):
        try:
            return pd.read_excel(
                io.BytesIO(
                    conteudo
                ),
                engine="xlrd",
                dtype=object,
            )

        except Exception:
            return pd.read_excel(
                io.BytesIO(
                    conteudo
                ),
                engine="openpyxl",
                dtype=object,
            )

    raise ValueError(
        "Formato não reconhecido. "
        "Envie CSV, XLS, XLSX ou XLSM."
    )


def valor_texto(
    linha,
    coluna,
    padrao="",
):
    if not coluna:
        return padrao

    valor = linha[coluna]

    try:
        if pd.isna(valor):
            return padrao
    except (TypeError, ValueError):
        pass

    texto = str(valor).strip()

    if texto:
        return texto

    return padrao


def intersecao_segundos(
    inicio_a,
    fim_a,
    inicio_b,
    fim_b,
):
    valores = [
        inicio_a,
        fim_a,
        inicio_b,
        fim_b,
    ]

    if any(
        pd.isna(valor)
        for valor in valores
    ):
        return 0.0

    inicio_intersecao = max(
        inicio_a,
        inicio_b,
    )

    fim_intersecao = min(
        fim_a,
        fim_b,
    )

    if fim_intersecao <= inicio_intersecao:
        return 0.0

    return float(
        (
            fim_intersecao
            - inicio_intersecao
        ).total_seconds()
    )
def reconstruir_setups_principal(df_bruto):
    df = preparar_dataframe(
        df_bruto
    )

    coluna_setup = localizar_coluna(
        df,
        ["setup"],
        obrigatoria=True,
    )

    coluna_descricao = localizar_coluna(
        df,
        [
            "descricao",
            "descrição",
        ],
        obrigatoria=True,
    )

    coluna_inicio = localizar_coluna(
        df,
        [
            "inicio",
            "início",
        ],
        obrigatoria=True,
    )

    coluna_fim = localizar_coluna(
        df,
        ["fim"],
        obrigatoria=True,
    )

    coluna_equipamento = localizar_coluna(
        df,
        [
            "equipamento",
            "maquina",
            "máquina",
        ],
        obrigatoria=True,
    )

    coluna_setor = localizar_coluna(
        df,
        ["setor"],
    )

    coluna_turno = localizar_coluna(
        df,
        ["turno"],
    )

    coluna_tempo = localizar_coluna(
        df,
        ["tempo"],
    )

    coluna_tempo_total = localizar_coluna(
        df,
        [
            "tempo total",
            "tempo_total",
        ],
    )

    df["__inicio"] = converter_datas(
        df[coluna_inicio]
    )

    df["__fim"] = converter_datas(
        df[coluna_fim]
    )

    df["__setup_norm"] = df[
        coluna_setup
    ].apply(
        normalizar_texto
    )

    df["__desc_norm"] = df[
        coluna_descricao
    ].apply(
        normalizar_texto
    )

    mascara_inicio_bloco = (
        df[coluna_setup].notna()
        & df[coluna_setup]
        .astype(str)
        .str.strip()
        .ne("")
    )

    inicios_blocos = df.index[
        mascara_inicio_bloco
    ].tolist()

    registros = []
    auditoria = []

    for posicao, indice_inicio in enumerate(
        inicios_blocos
    ):
        if posicao + 1 < len(inicios_blocos):
            indice_fim = (
                inicios_blocos[
                    posicao + 1
                ]
                - 1
            )
        else:
            indice_fim = len(df) - 1

        setup_normalizado = df.at[
            indice_inicio,
            "__setup_norm",
        ]

        if setup_normalizado != "setup matriz":
            continue

        bloco = df.loc[
            indice_inicio:indice_fim
        ].copy()

        mascara_aguardando = bloco[
            "__desc_norm"
        ].str.startswith(
            (
                "ag. trocador",
                "ag trocador",
                "aguardando trocador",
            ),
            na=False,
        )

        mascara_troca = bloco[
            "__desc_norm"
        ].str.startswith(
            "troca de molde",
            na=False,
        )

        linhas_aguardando = bloco[
            mascara_aguardando
        ].copy()

        linhas_troca = bloco[
            mascara_troca
        ].copy()

        inicio_setup = bloco[
            "__inicio"
        ].min()

        fim_setup = bloco[
            "__fim"
        ].max()

        if coluna_tempo_total:
            tempo_total_relatorio = (
                duracao_segundos(
                    bloco.iloc[0][
                        coluna_tempo_total
                    ]
                )
            )
        else:
            tempo_total_relatorio = np.nan

        if coluna_tempo:
            soma_tempo_linhas = soma_duracoes(
                bloco[coluna_tempo]
            )
        else:
            soma_tempo_linhas = np.nan

        if (
            pd.notna(inicio_setup)
            and pd.notna(fim_setup)
            and fim_setup >= inicio_setup
        ):
            tempo_total_datas = float(
                (
                    fim_setup
                    - inicio_setup
                ).total_seconds()
            )
        else:
            tempo_total_datas = np.nan

        if (
            pd.notna(
                tempo_total_relatorio
            )
            and tempo_total_relatorio > 0
        ):
            tempo_total = (
                tempo_total_relatorio
            )

            origem_tempo_total = (
                "Tempo total da primeira linha"
            )

        elif (
            pd.notna(
                soma_tempo_linhas
            )
            and soma_tempo_linhas > 0
        ):
            tempo_total = (
                soma_tempo_linhas
            )

            origem_tempo_total = (
                "Soma da coluna Tempo"
            )

        elif pd.notna(
            tempo_total_datas
        ):
            tempo_total = (
                tempo_total_datas
            )

            origem_tempo_total = (
                "Último fim menos primeiro início"
            )

        else:
            tempo_total = np.nan

            origem_tempo_total = (
                "Não disponível"
            )

        if not linhas_aguardando.empty:
            inicio_aguardando = (
                linhas_aguardando[
                    "__inicio"
                ].min()
            )

            fim_aguardando = (
                linhas_aguardando[
                    "__fim"
                ].max()
            )
        else:
            inicio_aguardando = pd.NaT
            fim_aguardando = pd.NaT

        if not linhas_troca.empty:
            inicio_troca = linhas_troca[
                "__inicio"
            ].min()

            fim_troca = linhas_troca[
                "__fim"
            ].max()
        else:
            inicio_troca = pd.NaT
            fim_troca = pd.NaT

        if (
            coluna_tempo
            and not linhas_aguardando.empty
        ):
            aguardando_segundos = (
                soma_duracoes(
                    linhas_aguardando[
                        coluna_tempo
                    ]
                )
            )
        else:
            aguardando_segundos = np.nan

        if (
            pd.isna(
                aguardando_segundos
            )
            and pd.notna(
                inicio_aguardando
            )
            and pd.notna(
                inicio_troca
            )
        ):
            aguardando_segundos = max(
                0.0,
                float(
                    (
                        inicio_troca
                        - inicio_aguardando
                    ).total_seconds()
                ),
            )

        if (
            coluna_tempo
            and not linhas_troca.empty
        ):
            troca_molde_segundos = (
                soma_duracoes(
                    linhas_troca[
                        coluna_tempo
                    ]
                )
            )
        else:
            troca_molde_segundos = np.nan

        if (
            pd.isna(
                troca_molde_segundos
            )
            and pd.notna(
                inicio_troca
            )
            and pd.notna(
                fim_troca
            )
        ):
            troca_molde_segundos = max(
                0.0,
                float(
                    (
                        fim_troca
                        - inicio_troca
                    ).total_seconds()
                ),
            )

        inicio_janela = (
            inicio_aguardando
        )

        tempo_deslocamento_segundos = 120

        if pd.notna(
            fim_troca
        ):
            fim_janela = (
                fim_troca
                + timedelta(
                    seconds=tempo_deslocamento_segundos
                )
            )
        else:
            fim_janela = pd.NaT

        if (
            pd.notna(inicio_janela)
            and pd.notna(fim_janela)
            and fim_janela > inicio_janela
        ):
            tempo_janela_segundos = float(
                (
                    fim_janela
                    - inicio_janela
                ).total_seconds()
            )
        else:
            tempo_janela_segundos = np.nan
        problemas = []

        if pd.isna(tempo_total):
            problemas.append(
                "tempo total inválido"
            )

        if pd.isna(
            inicio_aguardando
        ):
            problemas.append(
                "sem Aguardando Trocador"
            )

        if (
            pd.isna(inicio_troca)
            or pd.isna(fim_troca)
        ):
            problemas.append(
                "sem Troca de Molde"
            )

        if pd.isna(
            tempo_janela_segundos
        ):
            problemas.append(
                "janela não analisável"
            )

        setup_id = (
            f"SM-{len(registros) + 1:05d}"
        )

        equipamento = (
            normalizar_equipamento(
                bloco.iloc[0][
                    coluna_equipamento
                ]
            )
        )

        setor = valor_texto(
            bloco.iloc[0],
            coluna_setor,
            "Não informado",
        )

        turno = valor_texto(
            bloco.iloc[0],
            coluna_turno,
            "Não informado",
        )

        descricoes = sorted(
            {
                str(valor).strip()
                for valor in bloco[
                    coluna_descricao
                ].dropna()
                if str(valor).strip()
            }
        )

        registro = {
            coluna: np.nan
            for coluna in COLUNAS_BASE
        }

        registro.update(
            {
                "id_analise": setup_id,
                "setup_id": setup_id,
                "extra_id": "",
                "tipo_setup": "Principal",
                "equipamento": equipamento,
                "setor": setor,
                "turno": turno,
                "linha_inicial": (
                    indice_inicio + 2
                ),
                "linha_final": (
                    indice_fim + 2
                ),
                "inicio_setup": (
                    inicio_setup
                ),
                "fim_setup": (
                    fim_setup
                ),
                "tempo_total_segundos": (
                    tempo_total
                ),
                "tempo_total_formatado": (
                    formatar_duracao(
                        tempo_total
                    )
                ),
                "origem_tempo_total": (
                    origem_tempo_total
                ),
                "tempo_total_relatorio_segundos": (
                    tempo_total_relatorio
                ),
                "soma_tempo_linhas_segundos": (
                    soma_tempo_linhas
                ),
                "tempo_total_datas_segundos": (
                    tempo_total_datas
                ),
                "inicio_aguardando": (
                    inicio_aguardando
                ),
                "fim_aguardando": (
                    fim_aguardando
                ),
                "inicio_troca_molde": (
                    inicio_troca
                ),
                "fim_troca_molde": (
                    fim_troca
                ),
                "aguardando_segundos": (
                    aguardando_segundos
                ),
                "aguardando_formatado": (
                    formatar_duracao(
                        aguardando_segundos
                    )
                ),
                "troca_molde_segundos": (
                    troca_molde_segundos
                ),
                "troca_molde_formatada": (
                    formatar_duracao(
                        troca_molde_segundos
                    )
                ),
                "inicio_janela": (
                    inicio_janela
                ),
                "fim_janela": (
                    fim_janela
                ),
                "tempo_janela_segundos": (
                    tempo_janela_segundos
                ),
                "janela_simultaneidade_formatada": (
                    formatar_duracao(
                        tempo_janela_segundos
                    )
                ),
                "descricoes_bloco": (
                    " | ".join(
                        descricoes
                    )
                ),
                "descricao": (
                    "Setup Matriz"
                ),
                "codigo": "",
                "programa": "",
                "documento": "",
                "produto": "",
                "ferramental": "",
                "qualidade": (
                    "Completo"
                    if not problemas
                    else (
                        "Incompleto: "
                        + "; ".join(
                            problemas
                        )
                    )
                ),
                "participa_analise": (
                    pd.notna(
                        tempo_janela_segundos
                    )
                ),
                "setup_extra": False,
                "setup_simultaneo": False,
                "bloco_id": pd.NA,
                "bloco_critico": False,
                "setup_critico": False,
                "pico_maquinas_paradas": 0,
                "excedente_maximo": 0,
                "tempo_excesso_segundos": 0.0,
                "tempo_excesso_formatado": (
                    "00:00:00"
                ),
                "tempo_critico_individual_segundos": 0.0,
                "tempo_critico_individual_formatado": (
                    "00:00:00"
                ),
                "espera_coincidente_excesso_segundos": 0.0,
                "espera_coincidente_excesso_formatada": (
                    "00:00:00"
                ),
                "quantidade_criticos_bloco": 0,
                "equipamentos_criticos_bloco": "",
                "classificacao": (
                    "Dado não analisável"
                ),
                "justificativa": (
                    "Aguardando análise."
                ),
                "motivo_extra": "",
                "fonte_setup": (
                    "Relatório Principal"
                ),
            }
        )

        registros.append(
            registro
        )

        for descricao in descricoes:
            auditoria.append(
                {
                    "setup_id": setup_id,
                    "equipamento": equipamento,
                    "descricao": descricao,
                    "descricao_normalizada": (
                        normalizar_texto(
                            descricao
                        )
                    ),
                }
            )

    resultado = pd.DataFrame(
        registros,
        columns=COLUNAS_BASE,
    )

    auditoria_descricoes = (
        pd.DataFrame(
            auditoria
        )
    )

    return (
        resultado,
        auditoria_descricoes,
        df,
    )


def processar_disponibilidade(
    df_bruto,
):
    df = preparar_dataframe(
        df_bruto
    )

    coluna_descricao = localizar_coluna(
        df,
        [
            "descricao",
            "descrição",
        ],
        obrigatoria=True,
    )

    df[
        "__descricao_normalizada"
    ] = df[
        coluna_descricao
    ].apply(
        normalizar_texto
    )

    descricoes_setups_extras = {
        "setup matriz",
        "setup matriz reposicao e amostra",
    }

    disponibilidade_processada = df[
        df[
            "__descricao_normalizada"
        ].isin(
            descricoes_setups_extras
        )
    ].copy()

    disponibilidade_processada = (
        disponibilidade_processada.reset_index(
            drop=True
        )
    )

    return disponibilidade_processada
def identificar_setups_extras(
    disponibilidade,
    principal_processado,
):
    if disponibilidade.empty:
        return pd.DataFrame(
            columns=COLUNAS_BASE
        )

    coluna_equipamento_disponibilidade = (
        localizar_coluna(
            disponibilidade,
            ["equipamento"],
            obrigatoria=True,
        )
    )

    coluna_equipamento_principal = (
        localizar_coluna(
            principal_processado,
            ["equipamento"],
            obrigatoria=True,
        )
    )

    maquinas_principal = set(
        principal_processado[
            coluna_equipamento_principal
        ]
        .dropna()
        .apply(
            normalizar_equipamento
        )
        .tolist()
    )

    disponibilidade_extras = (
        disponibilidade.copy()
    )

    disponibilidade_extras[
        "__equip_norm"
    ] = disponibilidade_extras[
        coluna_equipamento_disponibilidade
    ].apply(
        normalizar_equipamento
    )

    disponibilidade_extras = (
        disponibilidade_extras[
            ~disponibilidade_extras[
                "__equip_norm"
            ].isin(
                maquinas_principal
            )
        ].copy()
    )

    aliases = {
        "setor": [
            "setor",
        ],
        "turno": [
            "turno",
        ],
        "inicio": [
            "inicio",
            "início",
        ],
        "fim": [
            "fim",
        ],
        "tempo": [
            "tempo",
        ],
        "descricao": [
            "descricao",
            "descrição",
        ],
        "codigo": [
            "codigo",
            "código",
        ],
        "programa": [
            "programa",
        ],
        "documento": [
            "documento",
        ],
        "produto": [
            "produto",
        ],
        "ferramental": [
            "ferramental",
        ],
    }

    colunas = {
        chave: localizar_coluna(
            disponibilidade_extras,
            alternativas,
        )
        for chave, alternativas
        in aliases.items()
    }

    registros = []

    for numero, (_, linha) in enumerate(
        disponibilidade_extras.iterrows(),
        start=1,
    ):
        extra_id = (
            f"EX-{numero:05d}"
        )

        if colunas["inicio"]:
            inicio_setup = pd.to_datetime(
                linha[
                    colunas["inicio"]
                ],
                errors="coerce",
                dayfirst=False,
            )
        else:
            inicio_setup = pd.NaT

        if colunas["fim"]:
            fim_setup = pd.to_datetime(
                linha[
                    colunas["fim"]
                ],
                errors="coerce",
                dayfirst=False,
            )
        else:
            fim_setup = pd.NaT

        if colunas["tempo"]:
            tempo_total = (
                duracao_segundos(
                    linha[
                        colunas["tempo"]
                    ]
                )
            )
        else:
            tempo_total = np.nan

        if (
            pd.isna(tempo_total)
            and pd.notna(inicio_setup)
            and pd.notna(fim_setup)
            and fim_setup >= inicio_setup
        ):
            tempo_total = float(
                (
                    fim_setup
                    - inicio_setup
                ).total_seconds()
            )

        if (
            pd.notna(inicio_setup)
            and pd.notna(fim_setup)
            and fim_setup > inicio_setup
        ):
            tempo_janela = float(
                (
                    fim_setup
                    - inicio_setup
                ).total_seconds()
            )
        else:
            tempo_janela = np.nan

        registro = {
            coluna: np.nan
            for coluna in COLUNAS_BASE
        }

        registro.update(
            {
                "id_analise": extra_id,
                "setup_id": "",
                "extra_id": extra_id,
                "tipo_setup": (
                    "Setup Extra"
                ),
                "equipamento": linha[
                    "__equip_norm"
                ],
                "setor": valor_texto(
                    linha,
                    colunas["setor"],
                    "Não informado",
                ),
                "turno": valor_texto(
                    linha,
                    colunas["turno"],
                    "Não informado",
                ),
                "linha_inicial": np.nan,
                "linha_final": np.nan,
                "inicio_setup": (
                    inicio_setup
                ),
                "fim_setup": (
                    fim_setup
                ),
                "tempo_total_segundos": (
                    tempo_total
                ),
                "tempo_total_formatado": (
                    formatar_duracao(
                        tempo_total
                    )
                ),
                "origem_tempo_total": (
                    "Tempo do relatório "
                    "de disponibilidade"
                ),
                "tempo_total_relatorio_segundos": (
                    tempo_total
                ),
                "soma_tempo_linhas_segundos": (
                    np.nan
                ),
                "tempo_total_datas_segundos": (
                    tempo_janela
                ),
                "inicio_aguardando": pd.NaT,
                "fim_aguardando": pd.NaT,
                "inicio_troca_molde": pd.NaT,
                "fim_troca_molde": pd.NaT,
                "aguardando_segundos": np.nan,
                "aguardando_formatado": (
                    "Não disponível"
                ),
                "troca_molde_segundos": np.nan,
                "troca_molde_formatada": (
                    "Não disponível"
                ),
                "inicio_janela": (
                    inicio_setup
                ),
                "fim_janela": (
                    fim_setup
                ),
                "tempo_janela_segundos": (
                    tempo_janela
                ),
                "janela_simultaneidade_formatada": (
                    formatar_duracao(
                        tempo_janela
                    )
                ),
                "descricoes_bloco": (
                    valor_texto(
                        linha,
                        colunas["descricao"],
                        "Setup Matriz",
                    )
                ),
                "descricao": valor_texto(
                    linha,
                    colunas["descricao"],
                    "Setup Matriz",
                ),
                "codigo": valor_texto(
                    linha,
                    colunas["codigo"],
                ),
                "programa": valor_texto(
                    linha,
                    colunas["programa"],
                ),
                "documento": valor_texto(
                    linha,
                    colunas["documento"],
                ),
                "produto": valor_texto(
                    linha,
                    colunas["produto"],
                ),
                "ferramental": valor_texto(
                    linha,
                    colunas["ferramental"],
                ),
                "qualidade": (
                    "Completo"
                    if pd.notna(
                        tempo_janela
                    )
                    else (
                        "Incompleto: "
                        "janela não analisável"
                    )
                ),
                "participa_analise": False,
                "setup_extra": True,
                "setup_simultaneo": False,
                "bloco_id": pd.NA,
                "bloco_critico": False,
                "setup_critico": False,
                "pico_maquinas_paradas": 0,
                "excedente_maximo": 0,
                "tempo_excesso_segundos": 0.0,
                "tempo_excesso_formatado": (
                    "00:00:00"
                ),
                "tempo_critico_individual_segundos": 0.0,
                "tempo_critico_individual_formatado": (
                    "00:00:00"
                ),
                "espera_coincidente_excesso_segundos": 0.0,
                "espera_coincidente_excesso_formatada": (
                    "00:00:00"
                ),
                "quantidade_criticos_bloco": 0,
                "equipamentos_criticos_bloco": "",
                "classificacao": (
                    "Setup Extra sem bloco principal"
                ),
                "justificativa": (
                    "O início não ocorreu dentro "
                    "de um bloco principal."
                ),
                "motivo_extra": (
                    "Equipamento ausente do "
                    "relatório principal"
                ),
                "fonte_setup": (
                    "Relatório de Disponibilidade"
                ),
            }
        )

        registros.append(
            registro
        )

    return pd.DataFrame(
        registros,
        columns=COLUNAS_BASE,
    )


def formar_blocos_principais(
    principal,
):
    colunas_blocos = [
        "bloco_base_id",
        "inicio_bloco",
        "fim_bloco",
    ]

    if principal.empty:
        return pd.DataFrame(
            columns=colunas_blocos
        )

    mascara_valida = (
        principal[
            "inicio_janela"
        ].notna()
        & principal[
            "fim_janela"
        ].notna()
        & (
            principal[
                "fim_janela"
            ]
            > principal[
                "inicio_janela"
            ]
        )
    )

    validos = principal[
        mascara_valida
    ].copy()

    validos = validos.sort_values(
        by=[
            "inicio_janela",
            "fim_janela",
            "id_analise",
        ],
        kind="mergesort",
    )

    blocos = []
    bloco_atual = []
    maior_fim = pd.NaT

    for indice, linha in validos.iterrows():
        inicio_atual = linha[
            "inicio_janela"
        ]

        fim_atual = linha[
            "fim_janela"
        ]

        if not bloco_atual:
            bloco_atual = [
                indice
            ]

            maior_fim = fim_atual
            continue

        if inicio_atual >= maior_fim:
            blocos.append(
                bloco_atual
            )

            bloco_atual = [
                indice
            ]

            maior_fim = fim_atual
        else:
            bloco_atual.append(
                indice
            )

            maior_fim = max(
                maior_fim,
                fim_atual,
            )

    if bloco_atual:
        blocos.append(
            bloco_atual
        )

    registros_blocos = []

    for numero, indices in enumerate(
        blocos,
        start=1,
    ):
        dados_bloco = principal.loc[
            indices
        ]

        registros_blocos.append(
            {
                "bloco_base_id": (
                    f"BP-{numero:04d}"
                ),
                "inicio_bloco": (
                    dados_bloco[
                        "inicio_janela"
                    ].min()
                ),
                "fim_bloco": (
                    dados_bloco[
                        "fim_janela"
                    ].max()
                ),
            }
        )

    return pd.DataFrame(
        registros_blocos,
        columns=colunas_blocos,
    )


def associar_extras(
    extras,
    blocos_principais,
):
    extras = extras.copy()

    extras[
        "bloco_base_id"
    ] = pd.NA

    if (
        extras.empty
        or blocos_principais.empty
    ):
        return extras

    for indice, extra in extras.iterrows():
        inicio_extra = extra[
            "inicio_setup"
        ]

        fim_extra = extra[
            "fim_setup"
        ]

        if (
            pd.isna(inicio_extra)
            or pd.isna(fim_extra)
            or fim_extra <= inicio_extra
        ):
            extras.at[
                indice,
                "participa_analise",
            ] = False

            extras.at[
                indice,
                "classificacao",
            ] = "Dado não analisável"

            extras.at[
                indice,
                "justificativa",
            ] = (
                "O Setup Extra não possui "
                "uma janela válida."
            )

            continue

        blocos_compativeis = (
            blocos_principais[
                (
                    blocos_principais[
                        "inicio_bloco"
                    ]
                    <= inicio_extra
                )
                & (
                    inicio_extra
                    < blocos_principais[
                        "fim_bloco"
                    ]
                )
            ]
        )

        if blocos_compativeis.empty:
            continue

        bloco_escolhido = (
            blocos_compativeis.iloc[0]
        )

        extras.at[
            indice,
            "bloco_base_id",
        ] = bloco_escolhido[
            "bloco_base_id"
        ]

        extras.at[
            indice,
            "participa_analise",
        ] = True

        extras.at[
            indice,
            "classificacao",
        ] = "Setup Extra associado"

        extras.at[
            indice,
            "justificativa",
        ] = (
            "O início ocorreu dentro "
            "de um bloco principal."
        )

    return extras

def analisar_simultaneidade_com_extras(
    principal,
    extras,
    capacidade,
):
    principal = principal.copy()
    extras = extras.copy()

    blocos_principais = formar_blocos_principais(
        principal
    )

    extras = associar_extras(
        extras,
        blocos_principais,
    )

    if extras.empty:
        extras_associados = pd.DataFrame(
            columns=COLUNAS_BASE
        )
    else:
        extras_associados = extras[
            extras[
                "participa_analise"
            ].eq(True)
        ].copy()

    base_combinada = pd.concat(
        [
            principal,
            extras_associados,
        ],
        ignore_index=True,
        sort=False,
    )

    valores_padrao = {
        "setup_simultaneo": False,
        "bloco_id": pd.NA,
        "bloco_critico": False,
        "setup_critico": False,
        "pico_maquinas_paradas": 0,
        "excedente_maximo": 0,
        "tempo_excesso_segundos": 0.0,
        "tempo_excesso_formatado": "00:00:00",
        "tempo_critico_individual_segundos": 0.0,
        "tempo_critico_individual_formatado": "00:00:00",
        "espera_coincidente_excesso_segundos": 0.0,
        "espera_coincidente_excesso_formatada": "00:00:00",
        "quantidade_criticos_bloco": 0,
        "equipamentos_criticos_bloco": "",
        "classificacao": "Dado não analisável",
        "justificativa": (
            "A janela operacional não "
            "está disponível."
        ),
    }

    for coluna, valor in valores_padrao.items():
        base_combinada[
            coluna
        ] = valor

    mascara_valida = (
        base_combinada[
            "inicio_janela"
        ].notna()
        & base_combinada[
            "fim_janela"
        ].notna()
        & (
            base_combinada[
                "fim_janela"
            ]
            > base_combinada[
                "inicio_janela"
            ]
        )
    )

    base_valida = base_combinada[
        mascara_valida
    ].copy()

    base_valida = base_valida.sort_values(
        by=[
            "inicio_janela",
            "fim_janela",
            "id_analise",
        ],
        kind="mergesort",
    )

    blocos_combinados = []
    bloco_atual = []
    maior_fim_bloco = pd.NaT

    for indice, linha in base_valida.iterrows():
        inicio_atual = linha[
            "inicio_janela"
        ]

        fim_atual = linha[
            "fim_janela"
        ]

        if not bloco_atual:
            bloco_atual = [
                indice
            ]

            maior_fim_bloco = fim_atual
            continue

        if inicio_atual >= maior_fim_bloco:
            blocos_combinados.append(
                bloco_atual
            )

            bloco_atual = [
                indice
            ]

            maior_fim_bloco = fim_atual
        else:
            bloco_atual.append(
                indice
            )

            maior_fim_bloco = max(
                maior_fim_bloco,
                fim_atual,
            )

    if bloco_atual:
        blocos_combinados.append(
            bloco_atual
        )

    registros_resumo_blocos = []
    registros_intervalos_excesso = []

    for numero_bloco, indices_bloco in enumerate(
        blocos_combinados,
        start=1,
    ):
        bloco_id = (
            f"B-{numero_bloco:04d}"
        )

        eventos_por_horario = {}

        for indice in indices_bloco:
            inicio = base_combinada.at[
                indice,
                "inicio_janela",
            ]

            fim = base_combinada.at[
                indice,
                "fim_janela",
            ]

            if inicio not in eventos_por_horario:
                eventos_por_horario[
                    inicio
                ] = {
                    "fim": [],
                    "inicio": [],
                }

            if fim not in eventos_por_horario:
                eventos_por_horario[
                    fim
                ] = {
                    "fim": [],
                    "inicio": [],
                }

            eventos_por_horario[
                inicio
            ][
                "inicio"
            ].append(
                indice
            )

            eventos_por_horario[
                fim
            ][
                "fim"
            ].append(
                indice
            )

        horarios_ordenados = sorted(
            eventos_por_horario.keys()
        )

        ativos = set()
        intervalos_atividade = []
        intervalos_excesso = []
        pico_maquinas = 0

        for posicao, horario in enumerate(
            horarios_ordenados
        ):
            eventos_horario = (
                eventos_por_horario[
                    horario
                ]
            )

            for indice_fim in eventos_horario[
                "fim"
            ]:
                ativos.discard(
                    indice_fim
                )

            for indice_inicio in eventos_horario[
                "inicio"
            ]:
                ativos.add(
                    indice_inicio
                )

            quantidade_ativos = len(
                ativos
            )

            pico_maquinas = max(
                pico_maquinas,
                quantidade_ativos,
            )

            existe_proximo_horario = (
                posicao + 1
                < len(horarios_ordenados)
            )

            if not existe_proximo_horario:
                continue

            proximo_horario = (
                horarios_ordenados[
                    posicao + 1
                ]
            )

            if proximo_horario <= horario:
                continue

            intervalo = {
                "inicio": horario,
                "fim": proximo_horario,
                "ativos": set(
                    ativos
                ),
                "quantidade_ativos": (
                    quantidade_ativos
                ),
            }

            intervalos_atividade.append(
                intervalo
            )

            if (
                quantidade_ativos
                > capacidade
            ):
                intervalos_excesso.append(
                    intervalo.copy()
                )

        bloco_critico = (
            pico_maquinas
            > capacidade
        )

        tempo_excesso_bloco = sum(
            float(
                (
                    intervalo["fim"]
                    - intervalo["inicio"]
                ).total_seconds()
            )
            for intervalo
            in intervalos_excesso
        )

        excedentes = [
            intervalo[
                "quantidade_ativos"
            ]
            - capacidade
            for intervalo
            in intervalos_excesso
        ]

        excedente_maximo = max(
            excedentes,
            default=0,
        )

        indices_criticos = set()

        for intervalo in intervalos_excesso:
            indices_criticos.update(
                intervalo[
                    "ativos"
                ]
            )

        if indices_criticos:
            equipamentos_criticos = sorted(
                base_combinada.loc[
                    list(
                        indices_criticos
                    ),
                    "equipamento",
                ]
                .dropna()
                .astype(str)
                .unique()
                .tolist()
            )
        else:
            equipamentos_criticos = []

        for intervalo in intervalos_excesso:
            indices_ativos = list(
                intervalo[
                    "ativos"
                ]
            )

            dados_ativos = (
                base_combinada.loc[
                    indices_ativos
                ]
            )

            equipamentos_ativos = sorted(
                dados_ativos[
                    "equipamento"
                ]
                .dropna()
                .astype(str)
                .tolist()
            )

            identificadores_ativos = sorted(
                dados_ativos[
                    "id_analise"
                ]
                .dropna()
                .astype(str)
                .tolist()
            )

            tipos_ativos = sorted(
                dados_ativos[
                    "tipo_setup"
                ]
                .dropna()
                .astype(str)
                .unique()
                .tolist()
            )

            duracao_intervalo = float(
                (
                    intervalo["fim"]
                    - intervalo["inicio"]
                ).total_seconds()
            )

            registros_intervalos_excesso.append(
                {
                    "bloco_id": (
                        bloco_id
                    ),
                    "inicio_excesso": (
                        intervalo[
                            "inicio"
                        ]
                    ),
                    "fim_excesso": (
                        intervalo[
                            "fim"
                        ]
                    ),
                    "tempo_excesso_segundos": (
                        duracao_intervalo
                    ),
                    "tempo_excesso_formatado": (
                        formatar_duracao(
                            duracao_intervalo
                        )
                    ),
                    "maquinas_ativas": (
                        intervalo[
                            "quantidade_ativos"
                        ]
                    ),
                    "capacidade_duplas": (
                        capacidade
                    ),
                    "excedente": (
                        intervalo[
                            "quantidade_ativos"
                        ]
                        - capacidade
                    ),
                    "equipamentos_ativos": (
                        ", ".join(
                            equipamentos_ativos
                        )
                    ),
                    "setups_ativos": (
                        ", ".join(
                            identificadores_ativos
                        )
                    ),
                    "tipos_ativos": (
                        ", ".join(
                            tipos_ativos
                        )
                    ),
                }
            )

        for indice in indices_bloco:
            linha = base_combinada.loc[
                indice
            ]

            setup_simultaneo = any(
                (
                    indice
                    in intervalo[
                        "ativos"
                    ]
                )
                and (
                    intervalo[
                        "quantidade_ativos"
                    ]
                    > 1
                )
                for intervalo
                in intervalos_atividade
            )

            setup_critico = (
                indice
                in indices_criticos
            )

            inicio_janela = linha[
                "inicio_janela"
            ]

            fim_janela = linha[
                "fim_janela"
            ]

            tempo_critico_individual = sum(
                intersecao_segundos(
                    inicio_janela,
                    fim_janela,
                    intervalo[
                        "inicio"
                    ],
                    intervalo[
                        "fim"
                    ],
                )
                for intervalo
                in intervalos_excesso
            )

            if bool(
                linha[
                    "setup_extra"
                ]
            ):
                espera_coincidente_excesso = 0.0
            else:
                espera_coincidente_excesso = sum(
                    intersecao_segundos(
                        linha[
                            "inicio_aguardando"
                        ],
                        linha[
                            "fim_aguardando"
                        ],
                        intervalo[
                            "inicio"
                        ],
                        intervalo[
                            "fim"
                        ],
                    )
                    for intervalo
                    in intervalos_excesso
                )

            if bool(
                linha[
                    "setup_extra"
                ]
            ):
                if setup_critico:
                    classificacao = (
                        "Setup Extra crítico por "
                        "excesso de capacidade"
                    )

                    justificativa = (
                        "A janela completa do Setup Extra "
                        "esteve ativa em pelo menos um "
                        "intervalo no qual a quantidade de "
                        "máquinas ativas superou a "
                        "capacidade de duplas."
                    )

                elif setup_simultaneo:
                    classificacao = (
                        "Setup Extra simultâneo"
                    )

                    justificativa = (
                        "A janela completa do Setup Extra "
                        "se sobrepôs a pelo menos um setup, "
                        "mas não participou de intervalo "
                        "acima da capacidade."
                    )

                else:
                    classificacao = (
                        "Setup Extra não crítico"
                    )

                    justificativa = (
                        "A janela completa do Setup Extra "
                        "não participou de intervalo acima "
                        "da capacidade."
                    )

            else:
                if setup_critico:
                    classificacao = (
                        "Setup crítico por excesso "
                        "de capacidade"
                    )

                    justificativa = (
                        "O setup principal esteve ativo "
                        "durante um intervalo acima da "
                        "capacidade de duplas."
                    )

                elif (
                    setup_simultaneo
                    and bloco_critico
                ):
                    classificacao = (
                        "Simultâneo fora do "
                        "intervalo de excesso"
                    )

                    justificativa = (
                        "Houve simultaneidade no bloco, "
                        "mas o setup não esteve ativo "
                        "durante o excesso."
                    )

                elif setup_simultaneo:
                    classificacao = (
                        "Simultâneo dentro "
                        "da capacidade"
                    )

                    justificativa = (
                        "Houve sobreposição efetiva, "
                        "mas a capacidade de duplas "
                        "não foi ultrapassada."
                    )

                else:
                    classificacao = (
                        "Sem simultaneidade"
                    )

                    justificativa = (
                        "A janela operacional não teve "
                        "sobreposição efetiva com outro "
                        "setup."
                    )

            atualizacoes = {
                "setup_simultaneo": (
                    setup_simultaneo
                ),
                "bloco_id": (
                    bloco_id
                ),
                "bloco_critico": (
                    bloco_critico
                ),
                "setup_critico": (
                    setup_critico
                ),
                "pico_maquinas_paradas": (
                    pico_maquinas
                ),
                "excedente_maximo": (
                    excedente_maximo
                ),
                "tempo_excesso_segundos": (
                    tempo_excesso_bloco
                ),
                "tempo_excesso_formatado": (
                    formatar_duracao(
                        tempo_excesso_bloco
                    )
                ),
                "tempo_critico_individual_segundos": (
                    tempo_critico_individual
                ),
                "tempo_critico_individual_formatado": (
                    formatar_duracao(
                        tempo_critico_individual
                    )
                ),
                "espera_coincidente_excesso_segundos": (
                    espera_coincidente_excesso
                ),
                "espera_coincidente_excesso_formatada": (
                    formatar_duracao(
                        espera_coincidente_excesso
                    )
                ),
                "quantidade_criticos_bloco": (
                    len(
                        indices_criticos
                    )
                ),
                "equipamentos_criticos_bloco": (
                    ", ".join(
                        equipamentos_criticos
                    )
                ),
                "classificacao": (
                    classificacao
                ),
                "justificativa": (
                    justificativa
                ),
            }

            for (
                coluna,
                valor,
            ) in atualizacoes.items():
                base_combinada.at[
                    indice,
                    coluna,
                ] = valor

        dados_bloco = base_combinada.loc[
            indices_bloco
        ].copy()

        dados_principais_bloco = (
            dados_bloco[
                dados_bloco[
                    "setup_extra"
                ].eq(False)
            ]
        )

        dados_extras_bloco = (
            dados_bloco[
                dados_bloco[
                    "setup_extra"
                ].eq(True)
            ]
        )

        equipamentos_bloco = sorted(
            dados_bloco[
                "equipamento"
            ]
            .dropna()
            .astype(str)
            .unique()
            .tolist()
        )

        registros_resumo_blocos.append(
            {
                "bloco_id": (
                    bloco_id
                ),
                "inicio_bloco": (
                    dados_bloco[
                        "inicio_janela"
                    ].min()
                ),
                "fim_bloco": (
                    dados_bloco[
                        "fim_janela"
                    ].max()
                ),
                "quantidade_setups": (
                    len(
                        dados_bloco
                    )
                ),
                "quantidade_principais": (
                    len(
                        dados_principais_bloco
                    )
                ),
                "quantidade_extras": (
                    len(
                        dados_extras_bloco
                    )
                ),
                "pico_maquinas_paradas": (
                    pico_maquinas
                ),
                "capacidade_duplas": (
                    capacidade
                ),
                "bloco_critico": (
                    bloco_critico
                ),
                "excedente_maximo": (
                    excedente_maximo
                ),
                "tempo_excesso_segundos": (
                    tempo_excesso_bloco
                ),
                "tempo_excesso_formatado": (
                    formatar_duracao(
                        tempo_excesso_bloco
                    )
                ),
                "quantidade_criticos": (
                    len(
                        indices_criticos
                    )
                ),
                "equipamentos": (
                    ", ".join(
                        equipamentos_bloco
                    )
                ),
                "equipamentos_criticos": (
                    ", ".join(
                        equipamentos_criticos
                    )
                ),
            }
        )

    resumo_blocos = pd.DataFrame(
        registros_resumo_blocos
    )

    intervalos_excesso = pd.DataFrame(
        registros_intervalos_excesso
    )

    if not base_combinada.empty:
        mapa_resultados = (
            base_combinada.set_index(
                "id_analise",
                drop=False,
            )
        )
    else:
        mapa_resultados = (
            pd.DataFrame()
        )

    colunas_atualizacao = list(
        valores_padrao.keys()
    )

    for indice, linha in principal.iterrows():
        identificador = linha[
            "id_analise"
        ]

        if (
            mapa_resultados.empty
            or identificador
            not in mapa_resultados.index
        ):
            continue

        dados_analisados = (
            mapa_resultados.loc[
                identificador
            ]
        )

        if isinstance(
            dados_analisados,
            pd.DataFrame,
        ):
            dados_analisados = (
                dados_analisados.iloc[0]
            )

        for coluna in colunas_atualizacao:
            principal.at[
                indice,
                coluna,
            ] = dados_analisados[
                coluna
            ]

    for indice, linha in extras.iterrows():
        identificador = linha[
            "id_analise"
        ]

        if (
            mapa_resultados.empty
            or identificador
            not in mapa_resultados.index
        ):
            continue

        dados_analisados = (
            mapa_resultados.loc[
                identificador
            ]
        )

        if isinstance(
            dados_analisados,
            pd.DataFrame,
        ):
            dados_analisados = (
                dados_analisados.iloc[0]
            )

        for coluna in colunas_atualizacao:
            extras.at[
                indice,
                coluna,
            ] = dados_analisados[
                coluna
            ]

    principal = principal.sort_values(
        by=[
            "inicio_setup",
            "fim_setup",
            "id_analise",
        ],
        kind="mergesort",
    ).reset_index(
        drop=True
    )

    if not extras.empty:
        extras = extras.sort_values(
            by=[
                "inicio_setup",
                "fim_setup",
                "id_analise",
            ],
            kind="mergesort",
        ).reset_index(
            drop=True
        )

    return (
        principal,
        extras,
        resumo_blocos,
        intervalos_excesso,
    )


def filtrar_dataframe(
    dataframe,
    data_inicial,
    data_final,
    setores,
    turnos,
    equipamentos,
    classificacoes=None,
):
    if dataframe.empty:
        return dataframe.copy()

    mascara = (
        dataframe[
            "inicio_setup"
        ]
        .dt.date
        .between(
            data_inicial,
            data_final,
        )
    )

    if setores:
        mascara = (
            mascara
            & dataframe[
                "setor"
            ].isin(
                setores
            )
        )

    if turnos:
        mascara = (
            mascara
            & dataframe[
                "turno"
            ].isin(
                turnos
            )
        )

    if equipamentos:
        mascara = (
            mascara
            & dataframe[
                "equipamento"
            ].isin(
                equipamentos
            )
        )

    if classificacoes:
        mascara = (
            mascara
            & dataframe[
                "classificacao"
            ].isin(
                classificacoes
            )
        )

    return dataframe[
        mascara
    ].copy()


def criar_resumo_periodo(
    principal,
    extras,
    agrupamento,
):
    frequencias = {
        "Hora": "h",
        "Dia": "D",
        "Mês": "M",
    }

    formatos = {
        "Hora": (
            "%d/%m/%Y %H:00"
        ),
        "Dia": (
            "%d/%m/%Y"
        ),
        "Mês": (
            "%m/%Y"
        ),
    }

    frequencia = frequencias[
        agrupamento
    ]

    formato = formatos[
        agrupamento
    ]

    partes_resumo = []

    configuracoes = [
        {
            "dataframe": principal,
            "prefixo": "Principais",
            "coluna_id": "setup_id",
        },
        {
            "dataframe": extras,
            "prefixo": "Extras",
            "coluna_id": "extra_id",
        },
    ]

    for configuracao in configuracoes:
        dataframe = configuracao[
            "dataframe"
        ]

        if dataframe.empty:
            continue

        dados = dataframe[
            dataframe[
                "inicio_setup"
            ].notna()
        ].copy()

        if dados.empty:
            continue

        dados[
            "periodo_dt"
        ] = (
            dados[
                "inicio_setup"
            ]
            .dt.to_period(
                frequencia
            )
            .dt.start_time
        )

        dados[
            "tipo_grafico"
        ] = np.where(
            dados[
                "setup_critico"
            ].eq(True),
            (
                configuracao[
                    "prefixo"
                ]
                + " críticos"
            ),
            (
                configuracao[
                    "prefixo"
                ]
                + " não críticos"
            ),
        )

        resumo_parcial = (
            dados.groupby(
                [
                    "periodo_dt",
                    "tipo_grafico",
                ],
                as_index=False,
            )
            .agg(
                quantidade=(
                    configuracao[
                        "coluna_id"
                    ],
                    "size",
                ),
                tempo_total_segundos=(
                    "tempo_total_segundos",
                    "sum",
                ),
                tempo_medio_segundos=(
                    "tempo_total_segundos",
                    "mean",
                ),
            )
        )

        partes_resumo.append(
            resumo_parcial
        )

    if partes_resumo:
        resumo = pd.concat(
            partes_resumo,
            ignore_index=True,
        )

        resumo[
            "periodo"
        ] = resumo[
            "periodo_dt"
        ].dt.strftime(
            formato
        )

        resumo = resumo.sort_values(
            by=[
                "periodo_dt",
                "tipo_grafico",
            ],
            kind="mergesort",
        ).reset_index(
            drop=True
        )
    else:
        resumo = pd.DataFrame(
            columns=[
                "periodo_dt",
                "periodo",
                "tipo_grafico",
                "quantidade",
                "tempo_total_segundos",
                "tempo_medio_segundos",
            ]
        )

    return resumo


def grafico_setups_periodo(
    resumo,
):
    figura = go.Figure()

    configuracoes_series = [
        {
            "tipo": (
                "Principais não críticos"
            ),
            "cor": CORES["azul"],
            "grupo": "principal",
        },
        {
            "tipo": (
                "Principais críticos"
            ),
            "cor": CORES["vermelho"],
            "grupo": "principal",
        },
        {
            "tipo": (
                "Extras não críticos"
            ),
            "cor": CORES["roxo_claro"],
            "grupo": "extras",
        },
        {
            "tipo": (
                "Extras críticos"
            ),
            "cor": CORES["roxo_escuro"],
            "grupo": "extras",
        },
    ]

    for configuracao in configuracoes_series:
        dados = resumo[
            resumo[
                "tipo_grafico"
            ].eq(
                configuracao[
                    "tipo"
                ]
            )
        ].copy()

        if dados.empty:
            customdata = None
        else:
            customdata = (
                np.column_stack(
                    [
                        dados[
                            "tempo_total_segundos"
                        ].apply(
                            formatar_duracao
                        ),
                        dados[
                            "tempo_medio_segundos"
                        ].apply(
                            formatar_duracao
                        ),
                    ]
                )
            )

        figura.add_bar(
            name=configuracao[
                "tipo"
            ],
            x=dados.get(
                "periodo",
                pd.Series(
                    dtype=str
                ),
            ),
            y=dados.get(
                "quantidade",
                pd.Series(
                    dtype=float
                ),
            ),
            marker_color=configuracao[
                "cor"
            ],
            offsetgroup=configuracao[
                "grupo"
            ],
            legendgroup=configuracao[
                "grupo"
            ],
            customdata=customdata,
            hovertemplate=(
                "Período: %{x}<br>"
                "Quantidade: %{y}<br>"
                "Tempo total: %{customdata[0]}<br>"
                "Tempo médio: %{customdata[1]}"
                "<extra>%{fullData.name}</extra>"
            ),
        )

    figura.update_layout(
        barmode="stack",
        bargap=0.30,
        bargroupgap=0.08,
        title={
            "text": (
                "Setups por Período"
            ),
            "x": 0.02,
            "xanchor": "left",
        },
        xaxis_title="Período",
        yaxis_title=(
            "Quantidade de setups"
        ),
        template="plotly_dark",
        height=560,
        legend={
            "orientation": "h",
            "x": 0,
            "y": 1.12,
        },
        margin={
            "l": 80,
            "r": 40,
            "t": 110,
            "b": 120,
        },
        font={
            "size": 13,
        },
    )

    figura.update_xaxes(
        tickangle=-30,
        automargin=True,
        showgrid=False,
    )

    figura.update_yaxes(
        automargin=True,
        rangemode="tozero",
        gridcolor=(
            "rgba(148,163,184,0.22)"
        ),
    )

    return figura


def grafico_linhas_tempos(
    principal,
    agrupamento,
):
    figura = go.Figure()

    if principal.empty:
        figura.update_layout(
            title=(
                "Comparativo de "
                "Tempos Médios"
            ),
            template="plotly_dark",
            height=500,
        )

        return figura

    frequencias = {
        "Hora": "h",
        "Dia": "D",
        "Mês": "M",
    }

    formatos = {
        "Hora": "%d/%m %H:00",
        "Dia": "%d/%m/%Y",
        "Mês": "%m/%Y",
    }

    frequencia = frequencias[
        agrupamento
    ]

    formato = formatos[
        agrupamento
    ]

    dados = principal[
        principal[
            "inicio_setup"
        ].notna()
    ].copy()

    dados[
        "periodo_dt"
    ] = (
        dados[
            "inicio_setup"
        ]
        .dt.to_period(
            frequencia
        )
        .dt.start_time
    )

    resumo_tempos = (
        dados.groupby(
            "periodo_dt",
            as_index=False,
        )
        .agg(
            tempo_total=(
                "tempo_total_segundos",
                "mean",
            ),
            aguardando=(
                "aguardando_segundos",
                "mean",
            ),
            troca=(
                "troca_molde_segundos",
                "mean",
            ),
        )
    )

    resumo_tempos[
        "periodo"
    ] = resumo_tempos[
        "periodo_dt"
    ].dt.strftime(
        formato
    )

    configuracoes_linhas = [
        {
            "coluna": "tempo_total",
            "nome": "Tempo total médio",
            "cor": CORES["azul"],
        },
        {
            "coluna": "aguardando",
            "nome": (
                "Média Aguardando Trocador"
            ),
            "cor": CORES["laranja"],
        },
        {
            "coluna": "troca",
            "nome": (
                "Média Troca de Molde"
            ),
            "cor": CORES["verde"],
        },
    ]

    for configuracao in configuracoes_linhas:
        coluna = configuracao[
            "coluna"
        ]

        customdata = resumo_tempos[
            coluna
        ].apply(
            formatar_duracao
        )

        figura.add_trace(
            go.Scatter(
                x=resumo_tempos[
                    "periodo"
                ],
                y=(
                    resumo_tempos[
                        coluna
                    ]
                    / 60
                ),
                mode="lines+markers",
                name=configuracao[
                    "nome"
                ],
                line={
                    "color": configuracao[
                        "cor"
                    ],
                    "width": 3,
                },
                marker={
                    "size": 8,
                },
                customdata=customdata,
                hovertemplate=(
                    "Período: %{x}<br>"
                    "%{fullData.name}: "
                    "%{customdata}"
                    "<extra></extra>"
                ),
            )
        )

    figura.update_layout(
        title={
            "text": (
                "Comparativo de Tempos "
                "Médios por Período"
            ),
            "x": 0.02,
            "xanchor": "left",
        },
        xaxis_title="Período",
        yaxis_title="Minutos",
        template="plotly_dark",
        height=500,
        hovermode="x unified",
        legend={
            "orientation": "h",
            "x": 0,
            "y": 1.12,
        },
        margin={
            "l": 80,
            "r": 40,
            "t": 110,
            "b": 100,
        },
        font={
            "size": 13,
        },
    )

    figura.update_xaxes(
        tickangle=-25,
        automargin=True,
        showgrid=False,
    )

    figura.update_yaxes(
        automargin=True,
        rangemode="tozero",
        gridcolor=(
            "rgba(148,163,184,0.22)"
        ),
    )

    return figura

def dataframe_exportavel(
    dataframe,
):
    if not isinstance(
        dataframe,
        pd.DataFrame,
    ):
        return pd.DataFrame()

    saida = dataframe.copy()

    for coluna in saida.columns:
        if pd.api.types.is_datetime64_any_dtype(
            saida[coluna]
        ):
            saida[coluna] = saida[
                coluna
            ].dt.strftime(
                "%d/%m/%Y %H:%M:%S"
            )

        elif saida[coluna].dtype == "object":
            saida[coluna] = saida[
                coluna
            ].apply(
                lambda valor: (
                    ", ".join(
                        map(
                            str,
                            valor,
                        )
                    )
                    if isinstance(
                        valor,
                        (
                            list,
                            set,
                            tuple,
                        ),
                    )
                    else valor
                )
            )

    return saida


def gerar_excel(
    principal,
    simultaneos,
    resumo_blocos,
    extras,
    configuracao,
):
    if not isinstance(
        principal,
        pd.DataFrame,
    ):
        principal = pd.DataFrame()

    if not isinstance(
        simultaneos,
        pd.DataFrame,
    ):
        simultaneos = pd.DataFrame()

    if not isinstance(
        resumo_blocos,
        pd.DataFrame,
    ):
        resumo_blocos = pd.DataFrame()

    if not isinstance(
        extras,
        pd.DataFrame,
    ):
        extras = pd.DataFrame()

    if not isinstance(
        configuracao,
        pd.DataFrame,
    ):
        configuracao = pd.DataFrame()

    if principal.empty:
        principais_criticos = (
            pd.DataFrame()
        )

        principais_simultaneos = (
            pd.DataFrame()
        )
    else:
        principais_criticos = principal[
            principal[
                "setup_critico"
            ].eq(True)
        ].copy()

        principais_simultaneos = principal[
            principal[
                "setup_simultaneo"
            ].eq(True)
        ].copy()

    if extras.empty:
        extras_criticos = pd.DataFrame()
        extras_simultaneos = pd.DataFrame()
        media_extras = np.nan
    else:
        extras_criticos = extras[
            extras[
                "setup_critico"
            ].eq(True)
        ].copy()

        extras_simultaneos = extras[
            extras[
                "setup_simultaneo"
            ].eq(True)
        ].copy()

        media_extras = extras[
            "tempo_total_segundos"
        ].dropna().mean()

    resumo_executivo = pd.DataFrame(
        {
            "Indicador": [
                "Setups principais",
                "Principais simultâneos",
                "Principais críticos",
                "Setups Extras",
                "Extras simultâneos",
                "Extras críticos",
                "Tempo médio dos Extras",
            ],
            "Valor": [
                len(principal),
                len(principais_simultaneos),
                len(principais_criticos),
                len(extras),
                len(extras_simultaneos),
                len(extras_criticos),
                formatar_duracao(
                    media_extras
                ),
            ],
        }
    )

    abas = {
        "Resumo Executivo": (
            resumo_executivo
        ),
        "Todos os Principais": (
            principal
        ),
        "Simultâneos": (
            simultaneos
        ),
        "Resumo dos Blocos": (
            resumo_blocos
        ),
        "Todos os Extras": (
            extras
        ),
        "Configuração": (
            configuracao
        ),
    }

    buffer_inicial = io.BytesIO()

    with pd.ExcelWriter(
        buffer_inicial,
        engine="openpyxl",
    ) as writer:
        for nome_aba, dados_aba in abas.items():
            dados_exportaveis = (
                dataframe_exportavel(
                    dados_aba
                )
            )

            if dados_exportaveis.empty:
                dados_exportaveis = pd.DataFrame(
                    {
                        "Informação": [
                            (
                                "Nenhum registro "
                                "disponível para os "
                                "filtros selecionados."
                            )
                        ]
                    }
                )

            dados_exportaveis.to_excel(
                writer,
                sheet_name=nome_aba[:31],
                index=False,
            )

    buffer_inicial.seek(0)

    workbook = load_workbook(
        buffer_inicial
    )

    for worksheet in workbook.worksheets:
        worksheet.sheet_state = "visible"
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

        worksheet.sheet_view.showGridLines = (
            False
        )

        for celula in worksheet[1]:
            celula.font = Font(
                bold=True,
                color="FFFFFF",
            )

            celula.fill = PatternFill(
                fill_type="solid",
                fgColor="1E3A8A",
            )

            celula.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        worksheet.row_dimensions[
            1
        ].height = 25

        for coluna in worksheet.columns:
            maior_tamanho = 0

            for celula in coluna:
                if celula.value is None:
                    tamanho = 0
                else:
                    tamanho = len(
                        str(
                            celula.value
                        )
                    )

                maior_tamanho = max(
                    maior_tamanho,
                    tamanho,
                )

            largura = min(
                55,
                max(
                    12,
                    maior_tamanho + 2,
                ),
            )

            letra_coluna = get_column_letter(
                coluna[0].column
            )

            worksheet.column_dimensions[
                letra_coluna
            ].width = largura

    if (
        "Resumo Executivo"
        in workbook.sheetnames
    ):
        workbook.active = (
            workbook.sheetnames.index(
                "Resumo Executivo"
            )
        )
    else:
        workbook.active = 0

    buffer_final = io.BytesIO()

    workbook.save(
        buffer_final
    )

    buffer_final.seek(0)

    return buffer_final.getvalue()


def aplicar_tema_escuro():
    css = """
    <style>
    .stApp {
        background: #0F172A;
        color: #E2E8F0;
    }

    .block-container {
        padding-top: 3.5rem;
        padding-bottom: 3rem;
        max-width: 1600px;
    }

    [data-testid="stSidebar"] {
        background: #111827;
        border-right: 1px solid #334155;
    }

    [data-testid="stSidebar"] * {
        color: #E2E8F0;
    }

    .hero {
        border-radius: 18px;
        padding: 26px 30px;
        color: #FFFFFF;
        background:
            linear-gradient(
                120deg,
                #1D4ED8,
                #0284C7 48%,
                #7C3AED
            );
        margin: 4px 0 24px 0;
        box-shadow:
            0 12px 30px
            rgba(37, 99, 235, 0.20);
        overflow: hidden;
    }

    .hero h1 {
        margin: 0;
        color: #FFFFFF;
        font-size: 2rem;
        line-height: 1.2;
    }

    .hero p {
        margin: 9px 0 0 0;
        color: #FFFFFF;
        opacity: 0.94;
        line-height: 1.5;
    }

    .section-title {
        color: #E2E8F0;
        font-size: 1.28rem;
        font-weight: 800;
        margin: 28px 0 14px 0;
        border-left: 5px solid #2563EB;
        padding-left: 11px;
    }

    .kpi-grid {
        display: grid;
        grid-template-columns:
            repeat(
                4,
                minmax(0, 1fr)
            );
        gap: 18px;
        margin-bottom: 22px;
        align-items: stretch;
    }

    .kpi-grid.extras {
        grid-template-columns:
            repeat(
                4,
                minmax(0, 1fr)
            );
    }

    .kpi-card {
        background: #111827;
        border: 1px solid #334155;
        border-radius: 16px;
        padding: 18px 20px;
        min-height: 128px;
        display: flex;
        flex-direction: column;
        justify-content: center;
        box-sizing: border-box;
        overflow: hidden;
    }

    .kpi-label {
        color: #94A3B8;
        font-size: 0.79rem;
        font-weight: 700;
        text-transform: uppercase;
        letter-spacing: 0.03em;
        line-height: 1.35;
        min-height: 42px;
        overflow-wrap: break-word;
    }

    .kpi-value {
        color: #F8FAFC;
        font-size: 1.68rem;
        font-weight: 800;
        line-height: 1.15;
        margin-top: 10px;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
        font-variant-numeric: tabular-nums;
    }

    [data-testid="stDataFrame"] {
        border: 1px solid #334155;
        border-radius: 12px;
        overflow: hidden;
    }

    div[data-testid="stExpander"] {
        background: #111827;
        border: 1px solid #334155;
        border-radius: 12px;
        overflow: hidden;
    }

    div[data-testid="stDownloadButton"] button {
        min-height: 48px;
        border: 0;
        border-radius: 10px;
        color: #FFFFFF;
        font-weight: 700;
        background:
            linear-gradient(
                90deg,
                #2563EB,
                #7C3AED
            );
    }

    div[data-testid="stDownloadButton"] button:hover {
        border: 0;
        color: #FFFFFF;
        filter: brightness(1.06);
    }

    @media only screen and (max-width: 1250px) {
        .kpi-grid,
        .kpi-grid.extras {
            grid-template-columns:
                repeat(
                    2,
                    minmax(0, 1fr)
                );
        }
    }

    @media only screen and (max-width: 700px) {
        .block-container {
            padding-top: 1.4rem;
        }

        .hero {
            border-radius: 14px;
            padding: 20px;
        }

        .hero h1 {
            font-size: 1.55rem;
        }

        .kpi-grid,
        .kpi-grid.extras {
            grid-template-columns:
                1fr;
        }

        .kpi-card {
            min-height: 105px;
        }

        .kpi-value {
            font-size: 1.45rem;
        }
    }
    </style>
    """

    st.markdown(
        css,
        unsafe_allow_html=True,
    )


def exibir_indicadores(
    indicadores,
    classe_adicional="",
):
    cartoes = []

    for rotulo, valor in indicadores:
        valor_textual = str(
            valor
        )

        cartao = (
            '<div class="kpi-card">'
            '<div class="kpi-label">'
            f"{rotulo}"
            "</div>"
            '<div class="kpi-value" '
            f'title="{valor_textual}">'
            f"{valor_textual}"
            "</div>"
            "</div>"
        )

        cartoes.append(
            cartao
        )

    html = (
        '<div class="kpi-grid '
        f'{classe_adicional}">'
        + "".join(cartoes)
        + "</div>"
    )

    st.markdown(
        html,
        unsafe_allow_html=True,
    )

def exibir_rodape_fixo():
    st.markdown(
        """
        <style>
        .rodape-fixo {
            position: fixed;
            left: 0;
            bottom: 0;
            width: 100%;
            z-index: 999999;
            padding: 9px 16px;
            text-align: center;
            color: #94A3B8;
            background: rgba(15, 23, 42, 0.97);
            border-top: 1px solid #334155;
            font-size: 13px;
            box-sizing: border-box;
            backdrop-filter: blur(8px);
        }

        .rodape-fixo strong {
            color: #F8FAFC;
        }

        .block-container {
            padding-bottom: 5rem !important;
        }

        @media only screen and (max-width: 700px) {
            .rodape-fixo {
                padding: 8px 10px;
                font-size: 12px;
            }
        }
        </style>

        <div class="rodape-fixo">
            Feito por:
            <strong>Daniel Viana de Carvalho</strong>
        </div>
        """,
        unsafe_allow_html=True,
    )

# Inicialização obrigatória
setups_extras = pd.DataFrame(
    columns=COLUNAS_BASE
)

setups_extras_filtrados = pd.DataFrame(
    columns=COLUNAS_BASE
)

disponibilidade_processada = pd.DataFrame()

resumo = pd.DataFrame()

resultado = pd.DataFrame(
    columns=COLUNAS_BASE
)

resultado_filtrado = pd.DataFrame(
    columns=COLUNAS_BASE
)

resumo_blocos = pd.DataFrame()

resumo_blocos_filtrado = pd.DataFrame()

intervalos_excesso = pd.DataFrame()

intervalos_excesso_filtrados = pd.DataFrame()

auditoria_descricoes = pd.DataFrame()

principal_processado = pd.DataFrame()

principal_bruto = pd.DataFrame()

disponibilidade_bruta = pd.DataFrame()


with st.sidebar:
    st.header(
        "⚙️ Configuração"
    )

    arquivo_principal = st.file_uploader(
        "Upload 1: Relatório principal",
        type=[
            "csv",
            "xls",
            "xlsx",
            "xlsm",
        ],
        key="arquivo_principal",
    )

    arquivo_disponibilidade = st.file_uploader(
        (
            "Upload 2: Relatório de "
            "perdas disponibilidade"
        ),
        type=[
            "csv",
            "xls",
            "xlsx",
            "xlsm",
        ],
        key="arquivo_disponibilidade",
    )

    quantidade_trocadores = st.number_input(
        "Quantidade de trocadores",
        min_value=0,
        value=8,
        step=1,
    )

    capacidade_duplas = (
        int(
            quantidade_trocadores
        )
        // 2
    )

    st.info(
        (
            f"Capacidade: {capacidade_duplas} "
            "dupla(s) completa(s)"
        )
    )


aplicar_tema_escuro()


st.markdown(
    """
    <div class="hero">
        <h1>Análise de Simultaneidade de Setups</h1>
        <p>
            Simultaneidade, capacidade,
            criticidade e Setups Extras.
        </p>
    </div>
    """,
    unsafe_allow_html=True,
)


if arquivo_principal is None:
    st.info(
        (
            "Envie o relatório principal "
            "para iniciar. O relatório de "
            "disponibilidade é opcional."
        )
    )

    st.stop()


try:
    with st.spinner(
        (
            "Processando os relatórios "
            "e calculando a simultaneidade..."
        )
    ):
        principal_bruto = ler_arquivo(
            arquivo_principal
        )

        (
            resultado_base,
            auditoria_descricoes,
            principal_processado,
        ) = reconstruir_setups_principal(
            principal_bruto
        )

        if (
            arquivo_disponibilidade
            is not None
        ):
            disponibilidade_bruta = (
                ler_arquivo(
                    arquivo_disponibilidade
                )
            )

            disponibilidade_processada = (
                processar_disponibilidade(
                    disponibilidade_bruta
                )
            )

            setups_extras = (
                identificar_setups_extras(
                    disponibilidade_processada,
                    principal_processado,
                )
            )
        else:
            setups_extras = pd.DataFrame(
                columns=COLUNAS_BASE
            )

        (
            resultado,
            setups_extras,
            resumo_blocos,
            intervalos_excesso,
        ) = analisar_simultaneidade_com_extras(
            resultado_base,
            setups_extras,
            capacidade_duplas,
        )

except Exception as erro:
    st.error(
        (
            "Falha no processamento: "
            f"{erro}"
        )
    )

    st.exception(
        erro
    )

    st.stop()


datas_disponiveis = []


if not resultado.empty:
    datas_disponiveis.extend(
        resultado[
            "inicio_setup"
        ]
        .dropna()
        .tolist()
    )


if not setups_extras.empty:
    datas_disponiveis.extend(
        setups_extras[
            "inicio_setup"
        ]
        .dropna()
        .tolist()
    )


if not datas_disponiveis:
    st.warning(
        (
            "Não foram encontradas datas "
            "válidas nos relatórios."
        )
    )

    st.stop()


data_minima = min(
    datas_disponiveis
).date()


data_maxima = max(
    datas_disponiveis
).date()


with st.sidebar:
    st.divider()

    st.subheader(
        "Filtros"
    )

    periodo_selecionado = st.date_input(
        "Período",
        value=(
            data_minima,
            data_maxima,
        ),
        min_value=data_minima,
        max_value=data_maxima,
    )

    if isinstance(
        periodo_selecionado,
        (
            list,
            tuple,
        ),
    ):
        if len(
            periodo_selecionado
        ) >= 2:
            data_inicial = (
                periodo_selecionado[0]
            )

            data_final = (
                periodo_selecionado[1]
            )
        elif len(
            periodo_selecionado
        ) == 1:
            data_inicial = (
                periodo_selecionado[0]
            )

            data_final = (
                periodo_selecionado[0]
            )
        else:
            data_inicial = data_minima
            data_final = data_maxima
    else:
        data_inicial = (
            periodo_selecionado
        )

        data_final = (
            periodo_selecionado
        )

    if data_inicial > data_final:
        data_inicial, data_final = (
            data_final,
            data_inicial,
        )

    agrupamento = st.selectbox(
        "Agrupamento",
        [
            "Hora",
            "Dia",
            "Mês",
        ],
        index=0,
    )

    setores_principais = set(
        resultado.get(
            "setor",
            pd.Series(
                dtype=str
            ),
        )
        .dropna()
        .astype(str)
        .tolist()
    )

    setores_extras = set(
        setups_extras.get(
            "setor",
            pd.Series(
                dtype=str
            ),
        )
        .dropna()
        .astype(str)
        .tolist()
    )

    setores_disponiveis = sorted(
        setores_principais
        | setores_extras
    )

    turnos_principais = set(
        resultado.get(
            "turno",
            pd.Series(
                dtype=str
            ),
        )
        .dropna()
        .astype(str)
        .tolist()
    )

    turnos_extras = set(
        setups_extras.get(
            "turno",
            pd.Series(
                dtype=str
            ),
        )
        .dropna()
        .astype(str)
        .tolist()
    )

    turnos_disponiveis = sorted(
        turnos_principais
        | turnos_extras
    )

    equipamentos_principais = set(
        resultado.get(
            "equipamento",
            pd.Series(
                dtype=str
            ),
        )
        .dropna()
        .astype(str)
        .tolist()
    )

    equipamentos_extras = set(
        setups_extras.get(
            "equipamento",
            pd.Series(
                dtype=str
            ),
        )
        .dropna()
        .astype(str)
        .tolist()
    )

    equipamentos_disponiveis = sorted(
        equipamentos_principais
        | equipamentos_extras
    )

    filtro_setor = st.multiselect(
        "Setor",
        setores_disponiveis,
    )

    filtro_turno = st.multiselect(
        "Turno",
        turnos_disponiveis,
    )

    filtro_equipamento = st.multiselect(
        "Equipamento",
        equipamentos_disponiveis,
    )

    filtro_classificacao = st.multiselect(
        "Classificação principal",
        CLASSIFICACOES_PRINCIPAL,
    )


resultado_filtrado = filtrar_dataframe(
    resultado,
    data_inicial,
    data_final,
    filtro_setor,
    filtro_turno,
    filtro_equipamento,
    filtro_classificacao,
)


setups_extras_filtrados = filtrar_dataframe(
    setups_extras,
    data_inicial,
    data_final,
    filtro_setor,
    filtro_turno,
    filtro_equipamento,
)


ids_blocos_filtrados = set()


if not resultado_filtrado.empty:
    ids_blocos_filtrados.update(
        resultado_filtrado[
            "bloco_id"
        ]
        .dropna()
        .astype(str)
        .tolist()
    )


if not setups_extras_filtrados.empty:
    ids_blocos_filtrados.update(
        setups_extras_filtrados[
            "bloco_id"
        ]
        .dropna()
        .astype(str)
        .tolist()
    )


if not resumo_blocos.empty:
    resumo_blocos_filtrado = resumo_blocos[
        resumo_blocos[
            "bloco_id"
        ]
        .astype(str)
        .isin(
            ids_blocos_filtrados
        )
    ].copy()
else:
    resumo_blocos_filtrado = pd.DataFrame()


if not intervalos_excesso.empty:
    intervalos_excesso_filtrados = (
        intervalos_excesso[
            intervalos_excesso[
                "bloco_id"
            ]
            .astype(str)
            .isin(
                ids_blocos_filtrados
            )
        ].copy()
    )
else:
    intervalos_excesso_filtrados = (
        pd.DataFrame()
    )


resumo = criar_resumo_periodo(
    resultado_filtrado,
    setups_extras_filtrados,
    agrupamento,
)


if resultado_filtrado.empty:
    principais_validos = (
        pd.DataFrame()
    )

    principais_simultaneos = (
        pd.DataFrame()
    )

    principais_criticos = (
        pd.DataFrame()
    )

    media_tempo_total = np.nan
    media_aguardando = np.nan
    media_troca = np.nan
else:
    principais_validos = (
        resultado_filtrado[
            resultado_filtrado[
                "tempo_total_segundos"
            ].notna()
        ].copy()
    )

    principais_simultaneos = (
        resultado_filtrado[
            resultado_filtrado[
                "setup_simultaneo"
            ].eq(True)
        ].copy()
    )

    principais_criticos = (
        resultado_filtrado[
            resultado_filtrado[
                "setup_critico"
            ].eq(True)
        ].copy()
    )

    media_tempo_total = (
        principais_validos[
            "tempo_total_segundos"
        ]
        .dropna()
        .mean()
        if not principais_validos.empty
        else np.nan
    )

    media_aguardando = (
        resultado_filtrado[
            "aguardando_segundos"
        ]
        .dropna()
        .mean()
    )

    media_troca = (
        resultado_filtrado[
            "troca_molde_segundos"
        ]
        .dropna()
        .mean()
    )


if setups_extras_filtrados.empty:
    extras_simultaneos = (
        pd.DataFrame()
    )

    extras_criticos = (
        pd.DataFrame()
    )

    media_extras = np.nan
else:
    extras_simultaneos = (
        setups_extras_filtrados[
            setups_extras_filtrados[
                "setup_simultaneo"
            ].eq(True)
        ].copy()
    )

    extras_criticos = (
        setups_extras_filtrados[
            setups_extras_filtrados[
                "setup_critico"
            ].eq(True)
        ].copy()
    )

    media_extras = (
        setups_extras_filtrados[
            "tempo_total_segundos"
        ]
        .dropna()
        .mean()
    )


todos_simultaneos = pd.concat(
    [
        principais_simultaneos,
        extras_simultaneos,
    ],
    ignore_index=True,
    sort=False,
)


todos_criticos = pd.concat(
    [
        principais_criticos,
        extras_criticos,
    ],
    ignore_index=True,
    sort=False,
)


if intervalos_excesso_filtrados.empty:
    tempo_acima_capacidade = 0.0
else:
    tempo_acima_capacidade = (
        intervalos_excesso_filtrados[
            "tempo_excesso_segundos"
        ]
        .dropna()
        .sum()
    )


if resumo_blocos_filtrado.empty:
    maior_pico = 0
else:
    maior_pico = (
        resumo_blocos_filtrado[
            "pico_maquinas_paradas"
        ]
        .dropna()
        .max()
    )

    if pd.isna(
        maior_pico
    ):
        maior_pico = 0


st.markdown(
    (
        '<div class="section-title">'
        "Indicadores principais"
        "</div>"
    ),
    unsafe_allow_html=True,
)


indicadores_principais = [
    (
        "Setups principais válidos",
        len(
            principais_validos
        ),
    ),
    (
        "Tempo total médio",
        formatar_duracao(
            media_tempo_total
        ),
    ),
    (
        "Setups simultâneos",
        len(
            todos_simultaneos
        ),
    ),
    (
        "Setups críticos",
        len(
            todos_criticos
        ),
    ),
    (
        "Média aguardando trocador",
        formatar_duracao(
            media_aguardando
        ),
    ),
    (
        "Média troca de molde",
        formatar_duracao(
            media_troca
        ),
    ),
    (
        "Tempo acima da capacidade",
        formatar_duracao(
            tempo_acima_capacidade
        ),
    ),
    (
        "Maior pico de máquinas",
        int(
            maior_pico
        ),
    ),
]


exibir_indicadores(
    indicadores_principais
)


st.markdown(
    (
        '<div class="section-title">'
        "Setups Extras"
        "</div>"
    ),
    unsafe_allow_html=True,
)


indicadores_extras = [
    (
        "Setups Extras",
        len(
            setups_extras_filtrados
        ),
    ),
    (
        "Extras simultâneos",
        len(
            extras_simultaneos
        ),
    ),
    (
        "Extras críticos",
        len(
            extras_criticos
        ),
    ),
    (
        "Tempo médio dos Extras",
        formatar_duracao(
            media_extras
        ),
    ),
]


exibir_indicadores(
    indicadores_extras,
    classe_adicional="extras",
)


st.markdown(
    (
        '<div class="section-title">'
        "Gráficos"
        "</div>"
    ),
    unsafe_allow_html=True,
)


figura_setups = grafico_setups_periodo(
    resumo
)


st.plotly_chart(
    figura_setups,
    use_container_width=True,
    key="grafico_setups_periodo",
)


st.markdown(
    (
        "<div style='height: 24px;'>"
        "</div>"
    ),
    unsafe_allow_html=True,
)


figura_tempos = grafico_linhas_tempos(
    resultado_filtrado,
    agrupamento,
)


st.plotly_chart(
    figura_tempos,
    use_container_width=True,
    key="grafico_linhas_tempos",
)


st.markdown(
    (
        '<div class="section-title">'
        "Tabelas"
        "</div>"
    ),
    unsafe_allow_html=True,
)


tabelas = {
    "Todos os setups principais": (
        resultado_filtrado
    ),
    "Simultâneos": (
        todos_simultaneos
    ),
    "Resumo dos blocos": (
        resumo_blocos_filtrado
    ),
    "Todos os setups extras": (
        setups_extras_filtrados
    ),
}


for titulo_tabela, dados_tabela in tabelas.items():
    with st.expander(
        (
            f"{titulo_tabela} "
            f"({len(dados_tabela)})"
        ),
        expanded=False,
    ):
        if dados_tabela.empty:
            st.info(
                (
                    "Nenhum registro disponível "
                    "para os filtros selecionados."
                )
            )
        else:
            st.dataframe(
                dados_tabela,
                use_container_width=True,
                hide_index=True,
            )


st.markdown(
    (
        '<div class="section-title">'
        "Exportação"
        "</div>"
    ),
    unsafe_allow_html=True,
)


configuracao = pd.DataFrame(
    {
        "Parâmetro": [
            "Quantidade de trocadores",
            "Capacidade de duplas",
            "Regra do Setup Extra",
            "Gerado em",
        ],
        "Valor": [
            int(
                quantidade_trocadores
            ),
            capacidade_duplas,
            (
                "A janela completa do Setup Extra "
                "participa da análise quando o início "
                "ocorre dentro de um bloco principal."
            ),
            datetime.now().strftime(
                "%d/%m/%Y %H:%M:%S"
            ),
        ],
    }
)


try:
    arquivo_excel = gerar_excel(
        resultado_filtrado,
        todos_simultaneos,
        resumo_blocos_filtrado,
        setups_extras_filtrados,
        configuracao,
    )

    nome_arquivo = (
        "analise_setups_"
        f"{datetime.now():%Y%m%d_%H%M%S}"
        ".xlsx"
    )

    st.download_button(
        label=(
            "📥 Exportar para Excel"
        ),
        data=arquivo_excel,
        file_name=nome_arquivo,
        mime=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        ),
        use_container_width=True,
    )

except Exception as erro_exportacao:
    st.error(
        (
            "Não foi possível gerar o Excel: "
            f"{erro_exportacao}"
        )
    )